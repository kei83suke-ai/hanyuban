import sys
import subprocess
import json
import os

# 必要なライブラリ（openpyxl）がない場合は自動インストール
try:
    import openpyxl
except ImportError:
    print("Excel読み込み用のライブラリをインストールしています...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

# パス設定 (お使いの環境に合わせて適宜修正してください)
# デフォルトではデスクトップの「中国語単語.xlsx」を探すように設定しています
EXCEL_PATH = "/Users/sonodakeisuke/Desktop/中国語単語.xlsx"
JS_PATH = "/Users/sonodakeisuke/Desktop/chinese_app_v2/js/data.js"

def extract_data():
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ エラー: Excelファイルが見つかりません: {EXCEL_PATH}")
        print("デスクトップに「中国語単語.xlsx」を作成するか、スクリプト内の EXCEL_PATH を修正してください。")
        return

    print(f"Excelファイル ({EXCEL_PATH}) を解析中...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    db = {
        "nouns": [],
        "verbs": [],
        "adjectives": [],
        "phrases": []
    }
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        name_lower = sheet_name.lower()
        
        # 期待する列構成: 
        # A:カテゴリ, B:中国語(簡体字), C:ピンイン, D:日本語, E:英語
        
        target_list = None
        if "名詞" in name_lower: target_list = db["nouns"]
        elif "動詞" in name_lower: target_list = db["verbs"]
        elif "形容詞" in name_lower: target_list = db["adjectives"]
        elif "挨拶" in name_lower or "定型句" in name_lower or "フレーズ" in name_lower: target_list = db["phrases"]
        
        if target_list is not None:
            for row in range(2, sheet.max_row + 1):
                cat = sheet.cell(row=row, column=1).value or "その他"
                zh = sheet.cell(row=row, column=2).value
                pinyin = sheet.cell(row=row, column=3).value or ""
                ja = sheet.cell(row=row, column=4).value
                en = sheet.cell(row=row, column=5).value or ""
                
                if zh and ja:
                    target_list.append({
                        "category": str(cat).strip(),
                        "zh": str(zh).strip(),
                        "pinyin": str(pinyin).strip(),
                        "ja": str(ja).strip(),
                        "en": str(en).strip()
                    })

    # ジャンル情報の定義
    genres_info = [
      { "id": "nouns", "label": "名詞", "enLabel": "Nouns", "icon": "🍎", "color": "#FF6B6B" },
      { "id": "verbs", "label": "動詞", "enLabel": "Verbs", "icon": "🏃", "color": "#4ECDC4" },
      { "id": "adjectives", "label": "形容詞", "enLabel": "Adjectives", "icon": "✨", "color": "#FFD93D" },
      { "id": "phrases", "label": "挨拶・定型句", "enLabel": "Phrases", "icon": "💬", "color": "#6C5CE7" }
    ]

    # JSデータの生成
    print("データをアプリ用に変換中...")
    js_content = f"const db = {json.dumps(db, ensure_ascii=False, indent=2)};\n\n"
    js_content += f"const genresInfo = {json.dumps(genres_info, ensure_ascii=False, indent=2)};"
    
    with open(JS_PATH, "w", encoding="utf-8") as f:
        f.write(js_content)
        
    print("="*50)
    print("🎉 抽出完了！中国語アプリのデータを更新しました。")
    print(f"👉 ブラウザで index.html をリロードして確認してください。")
    print("="*50)

if __name__ == "__main__":
    extract_data()
